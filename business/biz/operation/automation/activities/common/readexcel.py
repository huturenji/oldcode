import openpyxl
import os
from common.contants import DATA_DIR


# 用来创建对象保存用例数据的类
class CaseData:
    pass


class ReadExcel(object):

    def __init__(self, filename, sheet_name):
        self.filename = filename
        self.sheet_name = sheet_name

    def open(self):
        """打开工作薄，选择表单"""
        # 如果文件存在，则加载文件
        if os.path.exists(self.filename):
            self.workbook = openpyxl.load_workbook(self.filename)
            # 如果sheet页存在，就打开指定的sheet页
            if self.sheet_name in self.workbook.sheetnames:
                self.sheet = self.workbook[self.sheet_name]
            else:
                # sheet页不存在，创建该sheet页
                self.workbook.create_sheet(self.sheet_name)
                self.workbook.save(self.filename)
        else:
            # 文件不存在，则创建文件和sheet页
            wb_create = openpyxl.Workbook()
            print("self.sheet_name",type(self.sheet_name))
            wb_create.active.title = self.sheet_name
            wb_create.save(self.filename)

        self.workbook = openpyxl.load_workbook(self.filename)
        self.sheet = self.workbook[self.sheet_name]

    def close(self):
        """关闭工作薄对象，释放内存"""
        self.workbook.close()

    def read_data(self):
        """

        :return: 列表嵌套字典的格式
        """
        self.open()
        # 按行获取所有的格子
        rows = list(self.sheet.rows)
        # 获取表头行数据
        title = []
        for r in rows[0]:
            title.append(r.value)

        # 创建一个空列表 用来存放所有的用例数据
        cases = []
        # 遍历除了表头剩余的行
        for row in rows[1:]:
            # 创建一个空列表，用来存储该行的数据
            data = []
            # 再次遍历该行的每一个格子
            for r in row:
                # 将格子中的数据，添加到data中
                data.append(r.value)
            case = dict(zip(title, data))
            cases.append(case)
        # 关闭工作薄
        self.close()
        return cases

    def read_data_obj(self):
        """
        :return: 列表嵌套对象
        """
        self.open()
        # 按行获取所有的格子
        rows = list(self.sheet.rows)
        # 获取表头行数据
        title = []
        for r in rows[0]:
            title.append(r.value)

        # 创建一个空列表 用来存放所有的用例数据
        cases = []
        # 遍历除了表头剩余的行
        for row in rows[1:]:
            # 创建一个空列表，用来存储该行的数据
            data = []
            # 再次遍历该行的每一个格子
            for r in row:
                # 将格子中的数据，添加到data中
                data.append(r.value)
            # 将表头和数据打包转换为列表
            case = list(zip(title, data))
            # 创建一个对象用来保存该行用例数据
            case_obj = CaseData()
            # 遍历列表中该行用例数据，使用setattr设置为对象的属性和属性值
            for k, v in case:
                setattr(case_obj, k, v)
            # print(case_obj,case_obj.__dict__)
            # 将对象添加到cases这个列表中
            cases.append(case_obj)
        # 关闭工作薄
        self.close()
        # 返回cases(包含所有用例数据对象的列表)
        return cases

    def write_data(self, row=None, column=None, value=None, li=None):
        # 打开工作薄
        self.open()
        # 写入数据
        if row == None:
            self.sheet.append(li)
        else:
            self.sheet.cell(row=row, column=column, value=value)
        # 保存文件
        self.workbook.save(self.filename)
        # 关闭工作薄
        self.close()


if __name__ == '__main__':
    # excel = ReadExcel(r"C:\Users\lisa\Desktop\TestMedia\data\cases.xlsx", "getArticle")
    # data = excel.read_data()
    # print(data)
    # print(len(data))
    # data_file_name = os.path.join(DATA_DIR, "personalcenter.xlsx")
    # excel = ReadExcel(r"C:\Users\duli\Desktop\TestMedia\data\personalcenter.xlsx", "deleteRecord")

    # excel.write_data(row=1,column=1,value=2)
    # excel.write_data(li = [1,2,3])

    data_file_name = os.path.join(DATA_DIR, "center.xlsx")
    excel = ReadExcel(data_file_name, "deleteRecord")
    data = excel.read_data()
    print(data)
