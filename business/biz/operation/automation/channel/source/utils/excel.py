import openpyxl


class ExcelTools:
    def __init__(self, filename, sheet_name):
        self.filename = filename
        self.sheet_name = sheet_name
        self.workbook = None
        self.sheet = None

    def read(self):
        self.workbook = openpyxl.load_workbook(self.filename)
        self.sheet = self.workbook[self.sheet_name]
        excel_data = {
            "bankname": self.sheet.cell(row=2, column=4).value,
            "env": self.sheet.cell(row=3, column=4).value,
            "protocol": self.sheet.cell(row=4, column=4).value,
            "domain": self.sheet.cell(row=5, column=4).value,
            "port": self.sheet.cell(row=6, column=4).value,
            "mall_enable": self.sheet.cell(row=9, column=4).value,
            "travel_enable": self.sheet.cell(row=10, column=4).value,
            "media_enable": self.sheet.cell(row=11, column=4).value,
            "wx_pay": self.sheet.cell(row=12, column=4).value,
            "ali_pay": self.sheet.cell(row=13, column=4).value,
            "quick_pay": self.sheet.cell(row=14, column=4).value,
            "secondDomain": self.sheet.cell(row=7, column=4).value
        }
        return excel_data
